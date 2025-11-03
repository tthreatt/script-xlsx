import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os

input_dir = '/Users/tthreatt/Desktop/BCBS-SC'
output_file = 'merged_data_test.xlsx'

csv_files = [f for f in os.listdir(input_dir) if f.endswith('.csv')]

search_layout = [
    {
        "sheet": "dnpi_license_bcbs_sc_2025-10-0",
        "row": 3,
        "headers": [
            "first_name", "middle_name", "last_name", "organization_name", "monitored_product", "issuer", "license_type", "license_source", "multi_stata",
            "license_category", "verified_first_name", "verified_middle_name", "verified_last_name", "verified_org_name", "verified_license_issued", "verified_license_number",
            "verified_license_status", "verified_license_details", "verified_license_expiration", "calculated_license_status", "abms_moc_status", "abms_renewal_date",
            "abms_duration_type", "abms_reverification_date", "dea_schedules", "dea_license_state"
        ],
        "formula": "FILTER('dnpi_license_bcbs_sc_2025-10-0'!B2:AA69985, 'dnpi_license_bcbs_sc_2025-10-0'!A2:A69985 = $A$1, \"\")"
    },
    {
        "sheet": "dnpi_preclusion_BCBS_SC_2025-1",
        "row": 101,
        "headers": [
            "monitored_product",
            "first_name",
            "middle_name",
            "last_name",
            "organization_name",
            "dob",
            "ein",
            "address_lines",
            "city",
            "state",
            "postal",
            "general",
            "speciality",
            "business_name",
            "preclusion_npi",
            "preclusion_id",
            "excluded_date",
            "reinstated_date",
            "claim_rejected_date",
            "preclusion_first_name",
            "preclusion_last_name",
            "preclusion_middle_name",
            "preclusion_former_last_name"
        ],
        "formula": "FILTER('dnpi_preclusion_BCBS_SC_2025-1'!B2:X70000, 'dnpi_preclusion_BCBS_SC_2025-1'!A2:A70000 = $A$1, \"\")"
    },
    {
        "sheet": "dnpi_exclusion_BCBS_SC_2025-10",
        "row": 110,
        "headers": [
            "first_name", "middle_name", "last_name", "organization_name", "monitored_product", "akas", "cage", "dbas", "npis", "type", "upin", "source",
            "address_of_residence", "address_line_2_of_residence", "fax_of_residence", "zip_of_residence", "city_of_residence", "state_of_residence", "county_of_residence",
            "country_of_residence", "telephone_of_residence", "comments", "source_id", "speciality", "dob", "duns_numbers", "exclusion_code", "start_date", "exclusion_date",
            "exclusion_term", "reinstate_date", "delisted_date", "classification", "exclusion_notes", "prefix", "suffix", "exclusion_last", "exclusion_first", "exclusion_middle",
            "exclusion_former_last", "exclusion_license_number", "excluding_agency", "provider_number", "exclusion_organization_name"
        ],
        "formula": "FILTER('dnpi_exclusion_BCBS_SC_2025-10'!B2:AS70000, 'dnpi_exclusion_BCBS_SC_2025-10'!A2:A70000 = $A$1, \"\")"
    },
    {
        "sheet": "dnpi_opt_out_BCBS_SC_2025-10-2",
        "row": 120,
        "headers": [
            "monitored_product", "first_name", "middle_name", "last_name", "organization_name", "address_lines", "city", "state", "postal", "speciality",
            "opt_out_id", "optout_npi", "effective_date", "end_date", "optout_first_name", "optout_last_name", "optout_former_last_name", "eligible_to_order_and_refer"
        ],
        "formula": "FILTER('dnpi_opt_out_BCBS_SC_2025-10-2'!B2:S70000, 'dnpi_opt_out_BCBS_SC_2025-10-2'!A2:A70000 = $A$1, \"\")"
    },
    {
        "sheet": "ddnpi_npi_BCBS_SC_2025-10-23",
        "row": 130,
        "headers": [
            "first_name", "middle_name", "last_name", "organization_name", "monitored_product", "ein", "nppes_npi", "entity_type", "last_update_date",
            "replacement_npi", "nppes_last_name", "nppes_first_name", "nppes_credentials", "nppes_middle_name", "nppes_name_prefix", "nppes_name_suffix",
            "nppes_organization_name", "is_sole_proprietor", "provider_gender_code", "npi_deactivation_date", "npi_reactivation_date", "npi_deactivation_reason",
            "provider_enumeration_date", "mailing_address_of_residence", "mailing_address_line_2_of_residence", "mailing_fax_of_residence",
            "mailing_zip_of_residence", "mailing_city_of_residence", "mailing_state_of_residence", "mailing_county_of_residence", "mailing_country_of_residence",
            "mailing_telephone_of_residence", "practice_address_of_residence", "practice_address_line_2_of_residence", "practice_fax_of_residence",
            "practice_zip_of_residence", "practice_city_of_residence", "practice_state_of_residence", "practice_county_of_residence", "practice_country_of_residence",
            "practice_telephone_of_residence", "nppes_licenses_taxonomies"
        ],
        "formula": "FILTER('dnpi_npi_BCBS_SC_2025-10-23'!B2:AQ70000, 'dnpi_npi_BCBS_SC_2025-10-23'!A2:A70000 = $A$1, \"\")"
    }
]


# Step 1: Create data sheets with pandas, as in your current script
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for csv_file in csv_files:
        path = os.path.join(input_dir, csv_file)
        df = pd.read_csv(path, dtype=str)
        sheet_name = os.path.splitext(csv_file[:30])[0]
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    # Write a blank search sheet just as a placeholder
    pd.DataFrame().to_excel(writer, index=False, sheet_name='Search_Tab')

# Step 2: Open file with openpyxl to write formulas and headers
wb = openpyxl.load_workbook(output_file)
# Move Search_Tab to be the first sheet
if "Search_Tab" in wb.sheetnames and wb.sheetnames[-1] == "Search_Tab":
    wb.move_sheet("Search_Tab", offset=-len(wb.sheetnames)+1)

ws = wb['Search_Tab']

for block in search_layout:
    # place headers in row `block["row"]` starting at column 1 (A)
    for col_idx, header in enumerate(block["headers"], start=1):
        cell = ws.cell(row=block["row"], column=col_idx, value=header)
        cell.font = Font(bold=True)
    # place FILTER formula in first cell below headers (col 1, row+1)
    ws.cell(row=block["row"]+1, column=1, value=block["formula"])

wb.save(output_file)
