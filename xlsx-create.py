"""
BCBS CSV to Excel Converter with Dynamic Filtering
===================================================

This script processes multiple CSV files from a specified directory and consolidates
them into a single Excel workbook with the following features:
- Each CSV type gets its own sheet (license, NPI, exclusion, preclusion, opt-out, OFAC, SSDMF, missing)
- A 'Search_Tab' sheet with dynamic FILTER formulas to query data across all sheets
- Data validation dropdown for license issuer filtering

SETUP INSTRUCTIONS FOR DAN:
---------------------------
1. Ensure Python 3.x is installed on your system
   
2. Set up virtual environment (recommended):
   python3 -m venv venv
   source venv/bin/activate  # On Mac/Linux
   # OR
   venv\Scripts\activate  # On Windows
   
3. Install required packages:
   pip install -r requirements.txt
   
4. Configure the script (see CONFIGURATION SECTION below):
   - Update 'input_dir' to point to your CSV folder
   - Update 'output_file' to your desired output filename
   
5. Run the script:
   python xlsx-create.py
   
6. Find your output Excel file in the current directory

INPUT FILE REQUIREMENTS:
------------------------
CSV files must follow this naming convention:
- License data: dnpi_license_bcbs_sc_*.csv
- NPI data: dnpi_npi_bcbs_sc_*.csv
- Exclusion data: dnpi_exclusion_bcbs_sc_*.csv
- Preclusion data: dnpi_preclusion_bcbs_sc_*.csv
- Opt-out data: dnpi_opt_out_bcbs_sc_*.csv
- OFAC data: dnpi_ofac_bcbs_sc_*.csv
- SSDMF data: dnpi_ssdmf_bcbs_sc_*.csv
- Missing NPI: *npi_missing_license_creds*.csv

TROUBLESHOOTING:
----------------
- If you get "FileNotFoundError": Check that input_dir path is correct
- If you get "ModuleNotFoundError": Run pip install -r requirements.txt
- If Excel formulas don't work: Make sure Excel supports FILTER() function (Excel 365/2021+)

Author: Tony Threatt
Date: November 2024
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os
import tracemalloc
import psutil

tracemalloc.start()

def log_memory(location):
    """Helper function to log memory at any point"""
    current, peak = tracemalloc.get_traced_memory()
    available = psutil.virtual_memory().available / 1024 / 1024
    print(f"[{location}] Current: {current / 1024 / 1024:.1f} MB | Peak: {peak / 1024 / 1024:.1f} MB | Available RAM: {available:.1f} MB")

# ============================================================================
# CONFIGURATION SECTION - Update these paths for your environment
# ============================================================================

log_memory("Script start")

# Directory containing the input CSV files
input_dir = '/Users/tthreatt/Desktop/BCBS-SC'

# Output Excel filename (will be created in the current directory)
output_file = 'BCBS-251110-new.xlsx'

# Date string used in sheet names (format: YY-MM-DD)
today_str = datetime.now().strftime('%y-%m-%d')  # e.g., '25-11-03'

# Get list of all CSV files in the input directory
csv_files = [f for f in os.listdir(input_dir) if f.endswith('.csv')]
print("*** All files being processed: ***")
for fname in csv_files:
    print("-", fname)

for csv_file in csv_files:
    # Add the full path by joining input_dir with the filename
    csv_path = os.path.join(input_dir, csv_file)
    df = pd.read_csv(csv_path, chunksize=10000)
    
    log_memory(f"After reading {csv_file}")

# ============================================================================
# SEARCH LAYOUT CONFIGURATION
# ============================================================================
# This defines the structure of the Search_Tab sheet. Each block represents:
# - sheet_prefix: The logical name prefix for the sheet
# - row: The row number where this data type's section starts in Search_Tab
# - headers: Column headers that will be displayed in the Search_Tab
# - special_match: (optional) Custom matching logic for specific files
# ============================================================================

search_layout = [
    {
        "sheet_prefix": "license_",
        "row": 3,
        "headers": [
            "first_name", "middle_name", "last_name", "organization_name", "monitored_product", "issuer", "license_type", "license_source", "multi_stata",
            "license_category", "verified_first_name", "verified_middle_name", "verified_last_name", "verified_org_name", "verified_license_action", "verified_license_issued", "verified_license_number",
            "verified_license_status", "verified_license_details", "verified_license_expiration", "calculated_license_status", "board_action_text", "abms_moc_status", "abms_renewal_date",
            "abms_duration_type", "abms_reverification_date", "dea_schedules", "dea_license_state"
        ]
    },
    {
        "sheet_prefix": "npi_",
        "row": 101,
        "headers": [
            "first_name", "middle_name", "last_name", "organization_name", "monitored_product", "ein", "nppes_npi", "entity_type", "last_update_date",
            "replacement_npi", "nppes_last_name", "nppes_first_name", "nppes_credentials", "nppes_middle_name", "nppes_name_prefix", "nppes_name_suffix",
            "nppes_organization_name", "is_sole_proprietor", "provider_gender_code", "npi_deactivation_date", "npi_reactivation_date", "npi_deactivation_reason",
            "provider_enumeration_date", "mailing_address_of_residence", "mailing_address_line_2_of_residence", "mailing_fax_of_residence",
            "mailing_zip_of_residence", "mailing_city_of_residence", "mailing_state_of_residence", "mailing_county_of_residence", "mailing_country_of_residence",
            "mailing_telephone_of_residence", "practice_address_of_residence", "practice_address_line_2_of_residence", "practice_fax_of_residence",
            "practice_zip_of_residence", "practice_city_of_residence", "practice_state_of_residence", "practice_county_of_residence", "practice_country_of_residence",
            "practice_telephone_of_residence", "nppes_licenses_taxonomies"
        ]
    },
    {
        "sheet_prefix": "exclusion_",
        "row": 110,
        "headers": [
            "first_name", "middle_name", "last_name", "organization_name", "monitored_product", "akas", "cage", "dbas", "npis", "type", "upin", "source",
            "address_of_residence", "address_line_2_of_residence", "fax_of_residence", "zip_of_residence", "city_of_residence", "state_of_residence", "county_of_residence",
            "country_of_residence", "telephone_of_residence", "comments", "source_id", "speciality", "dob", "duns_numbers", "exclusion_code", "start_date", "exclusion_date",
            "exclusion_term", "reinstate_date", "delisted_date", "classification", "exclusion_notes", "prefix", "suffix", "exclusion_last", "exclusion_first", "exclusion_middle",
            "exclusion_former_last", "exclusion_license_number", "excluding_agency", "provider_number", "exclusion_organization_name"
        ]
    },
    {
        "sheet_prefix": "preclusion_",
        "row": 120,
        "headers": [
            "monitored_product",
            "first_name",
            "middle_name",
            "last_name",
            "organization_name",
            "dob",
            "ein",
            "address_lines",
            "city",
            "state",
            "postal",
            "general",
            "speciality",
            "business_name",
            "preclusion_npi",
            "preclusion_id",
            "excluded_date",
            "reinstated_date",
            "claim_rejected_date",
            "preclusion_first_name",
            "preclusion_last_name",
            "preclusion_middle_name",
            "preclusion_former_last_name"
        ]
    },
    {
        "sheet_prefix": "opt_out_",
        "row": 130,
        "headers": [
            "monitored_product", "first_name", "middle_name", "last_name", "organization_name", "address_lines", "city", "state", "postal", "speciality",
            "opt_out_id", "optout_npi", "effective_date", "end_date", "optout_first_name", "optout_last_name", "optout_former_last_name", "eligible_to_order_and_refer"
        ]
    },
    {
        "sheet_prefix": "ofac_",
        "row": 140,
        "headers": [
            "npi", "first_name", "middle_name", "last_name", "organization_name", "monitored_product", "monitored_start_date", "monitored_end_date",
            "ofac_organization_name", "ofac_first_name", "ofac_middle_name", "ofac_last_name", "ofac_suffix", "ofac_date_of_birth", "ofac_npi",
            "ofac_tin", "ofac_specialty", "ofac_country", "ofac_date", "ofac_reinstate_date", "ofac_terms", "ofac_comments"
        ]
    },
    {
        "sheet_prefix": "ssdmf_",
        "row": 150,
        "headers": [
            "npi", "first_name", "middle_name", "last_name", "ssn", "monitored_product", "monitored_start_date", "monitored_end_date", "status",
            "last_verify_time", "verification_result", "ssdmf_first", "ssdmf_last", "ssdmf_date_of_birth", "ssdmf_date_of_death"
        ]
    },
    {
        "sheet_prefix": "missing_",
        "special_match": "npi_missing_license_creds",
        "row": 160,
        "headers": ["npi"]
    }  
]

# ============================================================================
# FILE PREFIX MAPPING
# ============================================================================
# Maps the actual CSV file prefixes to logical sheet prefixes
# CSV files are named like: 'dnpi_license_bcbs_sc_YYYYMMDD.csv'
# They get converted to sheets named like: 'license_YY-MM-DD'
# ============================================================================

SHORT_PREFIX_MAP = {
    'dnpi_preclusion_bcbs_sc_': 'preclusion_',
    'dnpi_exclusion_bcbs_sc_': 'exclusion_',
    'dnpi_license_bcbs_sc_': 'license_',
    'dnpi_opt_out_bcbs_sc_': 'opt_out_',
    'dnpi_ssdmf_bcbs_sc_': 'ssdmf_',
    'dnpi_ofac_bcbs_sc_': 'ofac_',
    'dnpi_npi_bcbs_sc_': 'npi_'
}

# Invert the map to get logical prefix -> file prefix mapping
# Example: LOGICAL_TO_FILE_PREFIX['license_'] => 'dnpi_license_bcbs_sc_'
LOGICAL_TO_FILE_PREFIX = {v: k for k, v in SHORT_PREFIX_MAP.items()}

# ============================================================================
# STEP 1: PROCESS CSV FILES AND MAP TO SHEETS
# ============================================================================
# These dictionaries track:
# - sheet_data: Stores the DataFrame for each sheet
# - sheetname_lookup: Maps logical prefix to actual sheet name
# - sheet_rowcount: Tracks the number of rows in each sheet (for FILTER formulas)
# - sheet_colcount: Tracks the last column letter in each sheet (for FILTER formulas)
# ============================================================================

sheet_data = {}
sheetname_lookup = {}
sheet_rowcount = {}
sheet_colcount = {}

# Process each block in the search layout to find and load matching CSV files
for block in search_layout:
    prefix = block["sheet_prefix"]
    special = block.get("special_match", "")  # For special matching rules (e.g., missing NPI file)
    row = block["row"]
    
    # Get the actual file prefix to match (e.g., 'dnpi_license_bcbs_sc_')
    file_prefix = LOGICAL_TO_FILE_PREFIX.get(prefix, prefix)
    
    # Look through all CSV files to find matches
    for csv_file in csv_files:
        match = False
        
        # Special matching for files like 'npi_missing_license_creds'
        if special:
            match = special in csv_file
            if not match:
                continue
        else:
            # Standard matching: file must start with the prefix and NOT be the missing NPI file
            match = csv_file.startswith(file_prefix) and ('npi_missing_license_creds' not in csv_file)
            if not match:
                continue

        # Load the matching CSV file
        path = os.path.join(input_dir, csv_file)
        
        # Create unique sheet name (e.g., 'license_25-11-03')
        # If duplicate, append _1, _2, etc.
        base_sheet_name = f"{prefix}{today_str}"
        sheet_name = base_sheet_name
        c = 1
        while sheet_name in sheet_data:
            sheet_name = f"{base_sheet_name}_{c}"
            c += 1

        # Read CSV and store metadata
        df = pd.read_csv(path, dtype=str)
        sheet_data[sheet_name] = df
        
        # Store the last row number (data rows + header row)
        sheet_rowcount[sheet_name] = len(df) + 1
        
        # Store the last column letter (e.g., 'Z', 'AA', etc.)
        sheet_colcount[sheet_name] = get_column_letter(df.shape[1])
        
        # Map the logical prefix to the actual sheet name (for reference in formulas)
        if prefix not in sheetname_lookup:
            sheetname_lookup[prefix] = sheet_name

# Display mapping information for debugging
print("\nSheet mapping:")
for k, v in sheetname_lookup.items():
    print(f"  {k} --> {v}")

print("\nSheet names that will be created:")
for sheet_name in sheet_data.keys():
    print("-", sheet_name)

log_memory("After combining all data")

# ============================================================================
# STEP 2: CREATE EXCEL WORKBOOK WITH ALL SHEETS
# ============================================================================
# Write all DataFrames to Excel and create an empty Search_Tab sheet
# ============================================================================

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Write each CSV data to its own sheet
    for sheet_name, df in sheet_data.items():
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    
    # Create empty Search_Tab sheet (will be populated with formulas below)
    pd.DataFrame().to_excel(writer, index=False, sheet_name='Search_Tab')

# Access the workbook directly from the writer (NO reload needed!)
wb = writer.book
ws = wb['Search_Tab']

# Move Search_Tab to be the first sheet in the workbook
if "Search_Tab" in wb.sheetnames and wb.sheetnames[-1] == "Search_Tab":
    wb.move_sheet("Search_Tab", offset=-len(wb.sheetnames)+1)

log_memory("After creating data frame")
# ============================================================================
# STEP 3: ADD FILTER FORMULAS TO SEARCH_TAB
# ============================================================================
# For each data type, add:
# 1. Bold headers in the specified row
# 2. Dynamic FILTER formula in the row below
# 
# The FILTER formulas pull data from the corresponding sheet based on:
# - Cell A1: NPI/ID to search for (matches column A of data sheets)
# - Cell B1: (License sheet only) Issuer filter - "All" or specific issuer
# ============================================================================

for block in search_layout:
    prefix = block["sheet_prefix"]
    headers = block["headers"]
    row = block["row"]
    
    # Get the actual sheet name for this data type
    sheetname = sheetname_lookup.get(prefix)
    if not sheetname or sheetname not in wb.sheetnames:
        print(f"Warning: Sheet for prefix '{prefix}' not found in workbook. Skipping.")
        continue

    # Write bold headers for this section
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=header).font = Font(bold=True)
    
    # Get the actual data range for this sheet (from metadata stored earlier)
    filter_end_col = sheet_colcount.get(sheetname, 'A')  # Last column letter (e.g., 'Z')
    filter_end_row = sheet_rowcount.get(sheetname, 2)    # Last row number (e.g., 105)

    # Define the Excel range to filter (excludes first column which is the ID)
    excel_range = f'B2:{filter_end_col}{filter_end_row}'

    # Create FILTER formula - special logic for license sheet (dual filter)
    if prefix == "license_":
        # License sheet filters by both NPI (A1) and Issuer (B1)
        # If B1 = "All", show all issuers; otherwise filter by specific issuer
        formula = (
            f'FILTER(\'{sheetname}\'!{excel_range}, '
            f'(\'{sheetname}\'!A2:A{filter_end_row}=$A$1)*(($B$1="All")+'
            f'(\'{sheetname}\'!G2:G{filter_end_row}=$B$1)), "")'
        )
    else:
        # All other sheets: simple filter by NPI/ID in A1
        formula = (
            f'FILTER(\'{sheetname}\'!{excel_range}, \'{sheetname}\'!A2:A{filter_end_row}=$A$1, "")'
        )
    
    # Place the formula in the row below the headers
    ws.cell(row=row + 1, column=1, value=formula)
    
    log_memory("After creating filters")

# ============================================================================
# STEP 4: CREATE ISSUER DROPDOWN FOR LICENSE FILTERING
# ============================================================================
# Extract all unique issuers from the license sheet (column G / index 7)
# and create a data validation dropdown in cell B1 of Search_Tab
# 
# Two methods:
# 1. If list is short (<=255 chars), embed directly in validation formula
# 2. If list is long, create hidden 'IssuerList' sheet with values
# ============================================================================

license_sheetname = sheetname_lookup.get("license_")
if license_sheetname and license_sheetname in wb.sheetnames:
    search_ws = ws
    license_ws = wb[license_sheetname]
    
    # Set default value to "All"
    search_ws['B1'] = "All"
    
    # Extract unique issuers from column G (issuer column) of license sheet
    issuers = set()
    for row in license_ws.iter_rows(min_row=2, min_col=7, max_col=7, values_only=True):
        val = row[0]
        if val and val.strip():
            issuers.add(val.strip())
    
    # Create sorted list with "All" at the top
    issuer_list = sorted(issuers)
    issuer_list.insert(0, "All")
    
    # Method 1: Short list - embed in formula
    # Excel has a 255-character limit for inline list validation
    if sum(len(i) for i in issuer_list) + len(issuer_list) - 1 <= 255:
        dv = DataValidation(type="list", formula1='"{}"'.format(",".join(issuer_list)), allow_blank=True)
        search_ws.add_data_validation(dv)
        dv.add(search_ws["B1"])
    else:
        # Method 2: Long list - use hidden sheet
        if 'IssuerList' not in wb.sheetnames:
            wb.create_sheet('IssuerList')
        issuer_ws = wb['IssuerList']
        
        # Write issuers to hidden sheet
        for idx, issuer in enumerate(issuer_list, start=1):
            issuer_ws.cell(row=idx, column=1, value=issuer)
        
        # Create validation referencing hidden sheet
        dv = DataValidation(
            type="list",
            formula1=f'=IssuerList!$A$1:$A${len(issuer_list)}',
            allow_blank=True
        )
        search_ws.add_data_validation(dv)
        dv.add(search_ws["B1"])
        
        # Hide the IssuerList sheet
        issuer_ws.sheet_state = 'hidden'
else:
    print("Warning: License sheet not found. Issuer drop-down will be skipped.")

log_memory("After creating dropdown")

# ============================================================================
# STEP 5: SAVE THE WORKBOOK
# ============================================================================

wb.save(output_file)
log_memory("After writing XLSX")
print(f"\n✅ Success! Excel file created: {output_file}")
print("\nHow to use:")
print("1. Open the Excel file and go to the 'Search_Tab' sheet")
print("2. Enter an NPI/ID in cell A1")
print("3. (Optional) For licenses, select an issuer from the dropdown in cell B1")
print("4. The filtered results will appear automatically below each section")
print("\nGo and Filter!")